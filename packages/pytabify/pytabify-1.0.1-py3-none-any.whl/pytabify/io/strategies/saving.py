import json
import csv
from openpyxl import Workbook
from pytabify.io.interfaces.save import SavingStrategy
from pytabify.core.datatable import DataTable
from pytabify.utils.errors import FileWritingException

class JsonFileSavingStrategy(SavingStrategy):
    """JsonFileSavingStrategy"""
    @staticmethod
    def save(datatable: DataTable, path: str, encoding: str) -> list[dict[str, str]]:
        """save"""
        data = datatable.to_dict()
        with open(path, mode="w", encoding=encoding) as output_file:
            try:
                json.dump(data, output_file)
            except Exception as e:
                raise FileWritingException(
                    f"No fue posible guardar los datos en el json {path}. Mas detalles: {e}"
                ) from e

class CsvFileSavingStrategy(SavingStrategy):
    """CsvFileSavingStrategy"""
    @staticmethod
    def save(datatable: DataTable, path: str, encoding: str) -> list[dict[str, str]]:
        """save"""

        fieldnames = datatable.headers()
        sorted_fieldnames = list(
            map(
                lambda fieldname: fieldname.name,
                sorted(fieldnames, key=lambda field: field.index)
            )
        )
        data = datatable.to_dict()

        with open(path, mode="w", encoding=encoding, newline="") as output_file:
            try:
                writer = csv.DictWriter(output_file, fieldnames=sorted_fieldnames)
                writer.writeheader()
                writer.writerows(data)
            except Exception as e:
                raise FileWritingException(
                    f"No fue posible guardar los datos en el csv {path}. Mas detalles: {e}"
                ) from e

class XlsxFileSavingStrategy(SavingStrategy):
    """XlsxFileSavingStrategy"""
    @staticmethod
    def save(datatable: DataTable, path: str, encoding: str) -> list[dict[str, str]]:
        """save"""
        fieldnames = datatable.headers()
        sorted_fieldnames = list(
            map(
                lambda fieldname: fieldname.name,
                sorted(fieldnames, key=lambda field: field.index)
            )
        )
        data = datatable.to_dict()

        wb = Workbook()
        wb_sheet = wb.active

        for col_idx, header in enumerate(sorted_fieldnames, start=1):
            wb_sheet.cell(row=1, column=col_idx, value=header)

        for row_idx, fila_data in enumerate(data, start=2):
            for col_idx, header in enumerate(sorted_fieldnames, start=1):
                wb_sheet.cell(row=row_idx, column=col_idx, value=fila_data.get(header, ""))

        wb.save(path)
