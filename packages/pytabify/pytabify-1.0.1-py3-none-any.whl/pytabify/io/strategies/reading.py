import json
import csv
from openpyxl import load_workbook
from pytabify.io.interfaces.read import ReadingStrategy
from pytabify.utils.errors import  (
    FileReadingException,
    FileNotFoundException,
    SheetNameHasNotEmptyException,
    SheetNameDoesNotExistException
)

class JSONFileReadingStrategy(ReadingStrategy):
    """JsonFileReadingStrategy"""
    def read(self) -> list[dict[str, str]]:
        if not self._file_exists():
            raise FileNotFoundException(f"El archivo {self._path} NO Existe verifique la ruta.")

        with open(self._path, mode="r", encoding=self._encoding) as file:
            try:
                return json.load(file)
            except json.JSONDecodeError as exc:
                raise FileReadingException("Ocurrio un error al leer el archivo de datos json") from exc

class CSVFileReadingStrategy(ReadingStrategy):
    """CsvFileReadingStrategy"""
    def read(self) -> list[dict[str, str]]:
        if not self._file_exists():
            raise FileNotFoundError(f"El archivo {self._path} NO Existe verifique la ruta.")

        with open(self._path, mode="r", encoding=self._encoding) as file:
            try:
                reader = csv.DictReader(file)
                return list(reader)
            except Exception as exc:
                raise FileReadingException("Ocurrio un Error al leer el archivo de datos csv") from exc

class XLSXReadingStrategy(ReadingStrategy):
    """XlsxReadingStrategy"""
    def read(self) -> list[dict[str, str]]:
        if not self._file_exists():
            raise FileNotFoundError(f"El archivo {self._path} NO Existe verifique la ruta.")

        if self._sheet_name is None:
            raise SheetNameHasNotEmptyException("sheet_name debe ser definido")

        workbook = load_workbook(self._path)

        if self._sheet_name not in workbook.sheetnames:
            raise SheetNameDoesNotExistException(f"La hoja {self._sheet_name} no existe en el archivo")

        hoja = workbook[self._sheet_name]
        encabezados = [celda.value for celda in hoja[1]]

        data = []
        for fila in hoja.iter_rows(min_row=2, values_only=True):
            fila_dict = {}
            for indice, valor_celda in enumerate(fila):
                fila_dict[encabezados[indice]] = "" if valor_celda is None else str(valor_celda)
            data.append(fila_dict)
        return data
