from coursebox.core.info_paths import get_paths
import os
import re
import openpyxl
import numpy as np
# from line_profiler_pycharm import profile
import pandas as pd

INSTRUCTOR_ROW = 6
INSTRUCTOR_CHECKER_ROW = 31
STUDENT_ID_ROW = 3
STARS_ROW = 26

EVALUATION_ROW_START = 8
EVALUATION_ROW_END = 40
WEIGHT_ROW_START = 63
RANGE_MIN_COL = 5

DELTA_ALLOWED_ROW = 111 # The range of possible delta-values. Should be in an empty (new) row at bottom.


def parse_column_student_ids(v):
    sn = []
    if v is not None:
        if isinstance(v, int):
            v = str(v)
            v = "s" + ("0"*(6-len(v))) + v
        v = v.lower()
        o = re.findall(r'(s\d{6})', v)
        for g in o:
            sn.append(g)
    return sn


def parse_column_numpy(col, report_number, column):
    """ Parse a column assuming it is defined as a numpy array.
    This is the recommended method as it is much, much faster.
    """
    # ws = worksheet  # wb.worksheets[sheet]
    sn = []
    group_id = col[STUDENT_ID_ROW - 1-1] #).value

    # col = ['' if col[0] is np.NAN else x for x in col]

    for i in range(0, 3):
        v = col[i + STUDENT_ID_ROW-1]#, column=column).value
        sn += parse_column_student_ids(v)


    instructor = col[INSTRUCTOR_ROW-1]#, column=column).value
    approver = col[INSTRUCTOR_ROW+1-1]# , column=column).value

    if instructor:
        instructor = instructor.lower()
    if approver:
        approver = str(approver).lower()

    content = None
    comments = None
    appr_comments = None
    if report_number > 0 and sn:
        N = 38
        rarr = np.ndarray(shape=(N,1),dtype=np.dtype(object))
        for j in range(N):
            v = col[3 + STUDENT_ID_ROW+j-1]#, column=column).value
            rarr[j,0] = v
        content = rarr
        comments = col[EVALUATION_ROW_END+5-1]# , column=column).value
        appr_comments = col[EVALUATION_ROW_END+6-1]# , column=column).value

    cgroup = {'column_j': column, 'student_ids': sn, 'instructor': instructor, "approver": approver, 'content': content,
              "comments": comments, "approver_comments": appr_comments, 'missing_fields': [],
              'group_id': group_id}

    # Now, find errors... This involves first finding non-zero columns
    if report_number > 0 and sn:
        score = cgroup['content'][-3, 0]
        cgroup['score'] = score
        cgroup['pct'] = score2pct(score)

        # if report_number == 3: # this obviously needs fixing for next semester.
        #     raise Exception("No report number 3 anymore. ")
        #     I = []
        #     for i in range(42): # max number of evaluation fields (irrelevant)
        #         v1 = col[WEIGHT_ROW_START+i-1, RANGE_MIN_COL-1]# ).value
        #         v2 = col[WEIGHT_ROW_START+i-1, RANGE_MIN_COL+1-1]#).value
        #         if (v1 == -1 and v2 == 1) or (v1 == 0 and v2 == 4):
        #             I.append(i)
        #         if v1 == -1 and v2 == 1:
        #             # print("delta col")
        #             break
        #
        #     for i in I:
        #         w1 = worksheet.cell(row=WEIGHT_ROW_START + i, column=1).value
        #         w3_ = worksheet.cell(row=INSTRUCTOR_ROW + i+2, column=1).value # should agree with label in w1
        #         w2 = worksheet.cell(row=INSTRUCTOR_ROW + i+2, column=column).value
        #         if w2 == None:
        #             cgroup['missing_fields'].append( (i, w1) )
        #             if report_number < 3:
        #                 print("old report nr.")

    return cgroup



def parse_column(worksheet, report_number, column):
    """ This is the old method. It is very slow. Use the numpy-version above.
    """
    ws = worksheet  # wb.worksheets[sheet]
    sn = []
    group_id = ws.cell(row=STUDENT_ID_ROW - 1, column=column).value

    for i in range(0, 3):
        v = ws.cell(row=i + STUDENT_ID_ROW, column=column).value
        sn += parse_column_student_ids(v)

    instructor = ws.cell(row=INSTRUCTOR_ROW, column=column).value
    approver = ws.cell(row=INSTRUCTOR_ROW+1, column=column).value

    if instructor:
        instructor = instructor.lower()
    if approver:
        approver = str(approver).lower()

    content = None
    comments = None
    appr_comments = None
    if report_number > 0 and sn:
        N = 38
        rarr = np.ndarray(shape=(N,1),dtype=np.dtype(object))
        for j in range(N):
            v = ws.cell(row=3 + STUDENT_ID_ROW+j, column=column).value
            rarr[j,0] = v
        content = rarr
        comments = ws.cell(row=EVALUATION_ROW_END+5, column=column).value
        appr_comments = ws.cell(row=EVALUATION_ROW_END+6, column=column).value

    cgroup = {'column_j': column, 'student_ids': sn, 'instructor': instructor, "approver": approver, 'content': content,
              "comments": comments, "approver_comments": appr_comments, 'missing_fields': [],
              'group_id': group_id}

    # Now, find errors... This involves first finding non-zero columns
    if report_number > 0 and sn:
        score = cgroup['content'][-3, 0]
        cgroup['score'] = score
        cgroup['pct'] = score2pct(score)

        if report_number == 99: # this obviously needs fixing for next semester.
            raise Exception("No report number 3 anymore. ")
            I = []
            for i in range(42): # max number of evaluation fields (irrelevant)
                v1 = worksheet.cell(row=WEIGHT_ROW_START+i, column=RANGE_MIN_COL).value
                v2 = worksheet.cell(row=WEIGHT_ROW_START+i, column=RANGE_MIN_COL+1).value
                if (v1 == -1 and v2 == 1) or (v1 == 0 and v2 == 4):
                    I.append(i)
                if v1 == -1 and v2 == 1:
                    # print("delta col")
                    break

            for i in I:
                w1 = worksheet.cell(row=WEIGHT_ROW_START + i, column=1).value
                w3_ = worksheet.cell(row=INSTRUCTOR_ROW + i+2, column=1).value # should agree with label in w1
                w2 = worksheet.cell(row=INSTRUCTOR_ROW + i+2, column=column).value
                if w2 == None:
                    cgroup['missing_fields'].append( (i, w1) )
                    if report_number < 3:
                        print("old report nr.")

    return cgroup

def score2pct(score):
    if score is None:
        return None
    if isinstance(score, str):
        return score
    else:
        pct = score / 4
        return pct

def get_output_file():
    out = get_paths()['collected_project_evaluations.xlsx']
    if not os.path.exists(os.path.dirname(out)):
        os.mkdir(os.path.dirname(out))
    return out

def get_groups_from_report(repn):
    cls = []
    out = get_output_file()
    print("> Loading student report scores from: %s" % out)
    wb = openpyxl.load_workbook(out, data_only=True)
    # Perhaps find non-empty cols (i.e. those with content)
    maximal_groups = []
    maximal_groups_students = []

    for i in range(2, wb.worksheets[repn].max_column + 1):
        cp = parse_column(wb.worksheets[repn], report_number=repn, column=i)
        if len(cp['student_ids']) == 0 or cp['group_id'] is None:
            continue
        cls.append(cp)
    return cls

def load_reports(sheet=1):
    """
    Tue 2023: This is a  new report loading function which will return a dictionary format. It is more convenient than the
    legacy methods that probably need to be refactored at some point.

    :param sheets:
    :return:
    """
    # Load the reports from the excel file.
    out = get_output_file()
    raise Exception()
    # print("> time elapsed", time.time() - t0)
    # maximal_groups = []
    # maximal_groups_students = []
    # for repn in sheets:
    #
    #     pass
    repn = sheet
    # for repn in sheets:
    cls = []
    sheet = pd.read_excel(out, sheet_name=repn, index_col=None, header=None)
    sheet = sheet.fillna('')
    sheet = sheet.to_numpy()
    # to_numpy()
    for i in range(1, sheet.shape[1]):

        # for i in range(2, wb.worksheets[repn].max_column + 1):
        # print(i, wb.worksheets[repn].max_column)
        # s = pd.read_excel(out, sheet_name=1)
        cp = parse_column_numpy(sheet[:, i], report_number=repn, column=i)
        # cp = parse_column(wb.worksheets[repn], report_number=repn, column=i)
        if not cp['student_ids']:
            break
        cls.append(cp)

    rs = {}
    for g in cls:
        students = ''

        for sid in g['student_ids']:
            student = students.get(sid, None)
            if student is None:
                if repn > 0:  # don't care about project 0 (group registration)
                    print("Bad error: Student id %s not found. report evaluation malformed?" % sid)
            else:
                # student = student.pop()
                student['reports'][repn] = g
                if sid not in maximal_groups_students:
                    maximal_groups.append(g)
                    maximal_groups_students += g['student_ids']
    print("> time elapsed", time.time() - t0)

    pass

# @profile
def populate_student_report_results(students, verbose=False):
    # take students (list-of-dicts in the info format) and assign them the results from the reports.
    out = get_output_file()
    import time
    t0 = time.time()
    if verbose:
        print(f"> Loading student report scores from: {out}")
    if not os.path.exists(out):
        return students, []

    for k in students:
        students[k]['reports'] = {i: None for i in range(4)}
    import pandas as pd

    # wb = openpyxl.load_workbook(out, data_only=True, read_only=True)
    # Perhaps find non-empty cols (i.e. those with content)
    if verbose:
        print("> time elapsed", time.time() - t0)

    maximal_groups = []
    maximal_groups_students = []

    for repn in range(3, -1, -1):
        cls = []
        sheet = pd.read_excel(out, sheet_name=repn, index_col=None, header=None)
        sheet = sheet.fillna('')
        sheet = sheet.to_numpy()
        # to_numpy()
        for i in range(1,sheet.shape[1]):

            # for i in range(2, wb.worksheets[repn].max_column + 1):
            # print(i, wb.worksheets[repn].max_column)
            # s = pd.read_excel(out, sheet_name=1)

            cp = parse_column_numpy(sheet[:,i], report_number=repn, column=i)
            if repn == 1 and 1 > 2:
                # info['students']['s225785']['reports'][1]
                pass

            # cp = parse_column(wb.worksheets[repn], report_number=repn, column=i)
            if not cp['student_ids']:
                break
            cls.append(cp)

        for g in cls:
            for sid in g['student_ids']:
                student = students.get(sid, None)
                if student is None:
                    if repn > 0:  # don't care about project 0 (group registration)
                        print("Bad error: Student id %s not found. report evaluation malformed?"%sid)
                else:
                    # student = student.pop()
                    student['reports'][repn] = g
                    if sid not in maximal_groups_students:
                        maximal_groups.append(g)
                        maximal_groups_students += g['student_ids']
    if verbose:
        print("> time elapsed", time.time() -t0)
    return students, maximal_groups