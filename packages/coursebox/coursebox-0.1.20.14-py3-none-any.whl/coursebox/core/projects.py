import re
import tempfile
import tika
import os
import shutil
import openpyxl
import numpy as np
import itertools
import math
import glob
from tika import parser
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
import matplotlib.pyplot as plt
import langdetect
import xlwings as xw
from coursebox.core.projects_info import get_output_file, INSTRUCTOR_ROW, STUDENT_ID_ROW, parse_column
from coursebox.core.projects_info import EVALUATION_ROW_END, EVALUATION_ROW_START, WEIGHT_ROW_START, RANGE_MIN_COL, DELTA_ALLOWED_ROW
from coursebox.core.info import get_paths, class_information, semester_id
from coursebox.core import projects_info
from coursebox.core.projects_plagiarism import plagiarism_checker
from jinjafy.cache import cache_contains_dir, cache_update_dir
from jinjafy.plot.plot_helpers import get_colors
import time
from collections import defaultdict
import zipfile
import pandas as pd
from slider.slide import recursive_tex_collect

def get_dirs(zf):
    zip = zipfile.ZipFile(zf)
    fls = list(set([os.path.dirname(x) for x in zip.namelist()]))
    fls = [f for f in fls if len(f) > 0]
    return fls

def fix_handins_fuckup(project_id=2):
    """ Handle the problem with multiple hand-ins in DTU learn. """
    paths = get_paths()
    from coursebox.core.info import class_information
    info = class_information()
    zf = paths['instructor_project_evaluations'] + f"/zip{project_id}.zip"

    tas = [i['shortname'] for i in info['instructors'] ]
    ta_links = {i['shortname']: i for i in info['instructors']}

    ta_reports = {}
    for ta in tas:
        fname = paths['instructor_project_evaluations']  + f"/project_{project_id}_{ta}.zip"
        for r in get_dirs(fname):
            if r in ta_reports:
                raise Exception
            ta_reports[r] = ta

    fls = get_dirs(zf)
    # fls = [f for f in zip.namelist() if not f.endswith("tml") and f.endswith("/")]
    d = defaultdict(lambda: [])
    for l in fls:
        # print(l)
        group_id = int(l.split("-")[2].strip().split(" ")[1])
        date = l.split("-")[-1].strip().split(" ")
        hm = date[-2]
        if len(hm) < 4:
            hm = "0" + hm
        hm = hm[:2] + ":" + hm[2:]
        date[-2] = hm
        date = " ".join(date)
        import datetime
        date_time_obj = datetime.datetime.strptime(date, '%d %B, %Y %I:%M %p')

        d[group_id].append( {'file': l, 'date': date_time_obj, 'ta': ta_reports[l], 'group_id': group_id } )

    d = {k: v for k, v in d.items() if len(v)  > 1}

    ta_do_not = defaultdict(lambda: [])
    ta_do = defaultdict(lambda: [])
    for group_id, reports in d.items():
        recent = max( r['date'] for r in reports)
        for r in reports:
            if r['date'] == recent:
                ta_do[r['ta']].append(r)
            else:
                ta_do_not[r['ta']].append(r)

    fname = paths['instructor_project_evaluations'] + "/do_not_evaluate.txt"
    with open(fname, 'w') as f:
        for ta, reports in ta_do_not.items():
            f.write(ta_links[ta ]['name'] + f" ({ta})" + "\n")
            for r in reports:
                f.write("DO NOT EVALUATE: " + r['file'] + "\n")
            f.write("\n")
            f.write("\n")
    with open(fname, 'r') as f:
        print(f.read())
    return ta_do, ta_do_not


def handle_projects(verbose=False, gather_main_xlsx_file=True, plagiarism_check=False):
    paths = get_paths()
    info = class_information()
    if info['freeze_report_evaluation']:
        print("> Report evaluations are FROZEN, meaning TA changes are no longer taken into account")
        print("> grades, etc. now relies on report resource file:")
        print(paths['collected_project_evaluations.xlsx'])
    if info['CE']:
        return
    if plagiarism_check:
        plagiarism_checker(paths, info)
        return

    instructor_path = paths['instructor_project_evaluations']
    cache_changed_xlsx_files = False
    if gather_main_xlsx_file:
        cache_base = paths['instructor_project_evaluations']
        if cache_contains_dir(cache_base,cache_base, pattern="*.xlsx") and os.path.exists(paths['collected_project_evaluations.xlsx']):
            pass
        else:
            cache_changed_xlsx_files = True
            gather_instructor_sheets(info)
            print("> Done gathering main .xlsx file from instructor .xlsx files")
            cache_update_dir(cache_base, cache_base, pattern="*.xlsx")
            info = class_information()
            id = 's225785'
            pid = 1
            print(info['students'][id]['reports'][pid])


    zip1 = instructor_path + "/zip1.zip"
    zip2 = instructor_path + "/zip2.zip"
    zip3 = instructor_path + "/zip3.zip"
    zips = [None, zip1, zip2, zip3]

    for j,zf in enumerate(zips):
        ifiles = get_instructor_xlsx_files(info, j)
        ex =  [os.path.exists(f) for (f,_) in ifiles]
        if any(ex) and not all(ex):
            raise Exception("Partial list of instructor files") # ensure there are either no files or all files exist
        instructor_files_exist = ex.pop()

        if instructor_files_exist:
            # if instructor files are there, we should do nothing
            continue
        else: # instructor files do not exist
            if j == 0:
                copy_populate_from_template(paths, info, sheet_number=j, zip_file=None)

            elif os.path.exists(zf):
                # make a copy of report template and populate it with groups obtained from previous report evaluation.
                # all_groups = get_all_reports_from_collected_xlsx_file()
                copy_populate_from_template(paths, info, sheet_number=j, zip_file=zf)
                # distribute_zip_content(info, sheet=j, zf_base=zf)
            else:
                print("When available, please move downloaded copy of all reports from campusnet to destination:")
                print(zf)
    info['students']['s225785']['reports'][1]
    mkboxplots(info['students'],paths)

    if cache_changed_xlsx_files:
        # Recompute nag files if instructor xlsx files have been changed
        compute_error_files(info, paths)
    nags = []
    fs = glob.glob(paths['instructor_project_evaluations'] + "/PARSE_ERRORS_*.txt")
    for f in fs:
        name = f.split("_").pop()[:-4]
        if name != "transfer":
            v = [i for i in info['instructors'] if i['shortname'] == name]
            with open(f, "r") as file:
                s = file.read()
                nlines = len(s.split("\n"))
            if len(v) == 0:
                raise Exception("Empty instructor list. what the ...")

            ins_email = v[0]['email']
            nags.append((name, ins_email, nlines))
    Nerrors = sum([n[-1] for n in nags])
    err = []
    serrors = None
    if Nerrors > 0:
        err.append("\n> %i errors. These instructors have parse errors (.txt)" % sum([n[-1] for n in nags]))
        err.append("; ".join([email for (_, email, _) in nags]))
        err.append(", ".join([name.capitalize() for (name, _, _) in nags]))
        err.append("TA/Errors: " + ", ".join(["%s:%i" % (name.capitalize(), lines) for (name, _, lines) in nags]))
        err.append("---")

        serrors = "\n".join(err)
        print(serrors)

    else:
        print("No parse errors found")
    return serrors




def compute_error_files(info, paths):
    print("Recomputing nag files...")
    ERRORS = dict()
    students = info['students']
    for repn in range(3, -1, -1):
        ifiles = get_instructor_xlsx_files(info, sheet=repn)
        all_groups = []
        for out, ins in ifiles:
            if not os.path.exists(out):
                continue

            if not ins in ERRORS:
                ERRORS[ins] = []

            wb = openpyxl.load_workbook(out, data_only=True)
            if len(wb.worksheets) > 1:
                es = os.path.basename(out) + "> Entire workbook is malformed. Workbook must only contain a single sheet. Fix ASAP; aborting further operations."
                ERRORS[ins].append(es)
                continue

            cls = []
            for i in range(2, wb.worksheets[0].max_column + 1):
                cp = projects_info.parse_column(wb.worksheets[0], report_number=repn, column=i)
                if not cp['student_ids']:
                    continue
                cls.append(cp)

            for g in cls:
                gins = g.get('instructor')
                err_label = "File: '%s', column with student ids: '%s'"%(os.path.basename(out), ', '.join(g['student_ids']))
                if gins != ins:
                    gins = "Null" if not gins else gins
                    es = err_label +"> Corrupted sheet. Instructor name: '"  + gins + "' not found. Should perhaps be: '" + ins+"'?"
                    print(es)
                    ERRORS[ins].append(es)
                    continue

                if repn >= 1 and g.get("score", None) == None and len(g['student_ids'])>0:
                    es = err_label + "> Report not scored."
                    ERRORS[ins].append(es)

                if repn >= 1 and len(g.get('student_ids', [])) > 0:
                    if g.get('score', 0) is None:
                        es = err_label + f"> Report does not have a score. The sheet may have been used incorrectly or fields are missing"
                        ERRORS[ins].append(es)
                    elif g.get("score", 0) < 0 or g.get("score", 0) > 4:
                        es = err_label + f"> Report score is {g.get('score', 0)}. The report score has to be between 0 and 4; probably due to a too high value of 'Delta' in instructor sheet."
                        ERRORS[ins].append(es)

                if repn >= 1 and not g['comments'] and info['course_number'] != '02465':
                    es = err_label + "> Incomplete report evaluation (missing comments field)"
                    es += "Please fill out comments field in your excel sheet."
                    ERRORS[ins].append(es)


                if repn >= 1 and not g['approver_comments']  and info['course_number'] != '02465':
                    es = err_label + "> Incomplete report evaluation (you are missing the approver comments field; can simply be set to 'ok')."
                    ERRORS.get(g['approver'], []).append(es)

                if g['missing_fields']:
                    mf = g['missing_fields']
                    es = err_label + "> Incomplete report evaluation (missing required evaluation-scoring field(s): '%s')."%(', '.join( [s for _,s in mf] ) ,)
                    es += " Please fill out missing field in your excel sheet."
                    ERRORS[ins].append(es)

                for sid in g['student_ids']:
                    # student = [student for student in students if student['id'] == sid]
                    if repn >= 1 and sid not in students and False:
                        # I disabled this since
                        es = err_label + "> Student ID '%s' not found in registered students on campusnet. "%sid
                        es += " To fix this, check if student is on campusnet under 'list of participants'; if he/she IS on campusnet, email me about the problem. " \
                              + " Otherwise, simply delete extra students from sheet and send an email to all students in the group (the student likely dropped out of course)"

                        ERRORS[ins].append( es )
                        continue

                    bad = [ (i, s, g) for (i,s,g) in all_groups if s == "sid"]
                    if len(bad) > 0:
                        i = bad[0][0]
                        g2 = bad[0][2]

                        es = err_label + "> Duplicate student ids for student: %s. " % (sid, )
                        es += "Student is also found in sheet by instructor: " + i
                        es += " in group with students: " + (", ".join( g2['student_ids'])) + ". "
                        es += " Please ensure report only assigned to one instructor. "
                        ERRORS[ins].append(es)
                        continue

                    all_groups.append( (ins, sid, g) )

    ipath = paths['instructor_project_evaluations']
    for f in glob.glob(ipath +"/PARSE_ERRORS_*.txt" ):
        os.remove(f)

    for ins in ERRORS:
        if ERRORS[ins]:
            ss = '\n'.join(ERRORS[ins])
            with open(ipath +"/PARSE_ERRORS_" + ins + ".txt",'w') as f:
                f.write(ss)


def get_template():
    paths = get_paths()
    return paths['project_evaluations_template.xlsx']


def get_instructor_xlsx_files(info, sheet):
    ss = "groups" if sheet == 0 else "report_%i"%sheet
    xlsx = []
    ins_names = [ins['shortname'] for ins in info['instructors']]
    ins_names.append("transfer")

    paths = get_paths()
    instructor_path = paths['instructor_project_evaluations']

    for ins in ins_names:
        ns = instructor_path + "/02450_"+semester_id()+"_" + ss + "_" + ins+".xlsx"
        xlsx.append( (ns,ins) )
    return xlsx


def get_groups_from_learn_xslx_file(paths, sheet_number):
    fname = f"{paths['instructor_project_evaluations']}/groups{sheet_number}.xlsx"
    all_groups = []
    if os.path.exists(fname):
        # Reading from the groups{number}.xlsx group-id file exported from DTU learn. Note this file contains fuckups.
        dg = defaultdict(list)
        df = pd.read_excel(fname)
        for uname, group_id in zip(df['Username'], df['Project groups']):
            id = int(group_id.split(" ")[1])
            if len(uname) == 7 and uname[0] == 's':
                dg[id].append(uname)
            else:
                dg[id].append("DTU-LEARN-FUCKED-THIS-ID-UP-CHECK-ON-REPORT")

        all_groups = [{'group_id': id, 'student_ids': students} for id, students in dg.items()]
    return all_groups


def group_id_from_file(file):
    id = int(os.path.dirname(file).split(" - ")[1].split(" ")[1])
    return id

def search_projects(paths, sheet_number, patterns):
    zip_files = [paths['instructor_project_evaluations'] + "/zip%d.zip" % sheet_number]
    # print(zip_files)

    all_groups = []
    gps = defaultdict(list)
    for zip_file in zip_files:
        if os.path.exists(zip_file):
            tmpdir = tempfile.TemporaryDirectory()
            zipfile.ZipFile(zip_file).extractall(path=tmpdir.name)

            # Read from PDF files:
            pdfs = glob.glob(tmpdir.name + "/**/*.pdf", recursive=True)
            for pdf in pdfs:
                pdf_parsed = tika.parser.from_file(pdf)
                id =group_id_from_file(pdf) # int(os.path.dirname(pdf).split(" - ")[1].split(" ")[1])
                if pdf_parsed['content'] is None:
                    students = []
                    print("> Finding student ID. Warning: The pdf file", pdf, "appers to have no text content.")
                else:
                    students = re.findall(r's\d\d\d\d\d\d', pdf_parsed['content'], flags=re.IGNORECASE)
                gps[id] += students

            # Collect from .tex files:

            # recursive_tex_collect()
            texs = glob.glob(tmpdir.name + "/**/*.tex", recursive=True)
            for tex in texs:
                id = group_id_from_file(tex)
                tex_parsed = recursive_tex_collect(tex)
                tex_parsed = "\n".join([(l[:l.find("%")] if "%" in l else l) for l in tex_parsed.splitlines()])
                students = re.findall(r's\d\d\d\d\d\d', tex_parsed, flags=re.IGNORECASE)
                gps[id] += students

    for id, students in gps.items():
        all_groups.append({'group_id': id, 'student_ids': list(set(students))})
    return all_groups


def unpack_zip_file_recursively(zip_file, destination_dir, remove_zipfiles=False):
    """
    Unpack the zip_file (extension: .zip) to the given directory.

    If the folders in the zip file contains other zip/files, these are unpacked recursively.
    """
    # Unpack zip file recursively and flatten it.
    zipfile.ZipFile(zip_file).extractall(path=destination_dir)
    ls = glob.glob(destination_dir + "/*")
    for f in ls:
        if os.path.isdir(f):
            zipfiles = glob.glob(f + "/*.zip")
            for zp in zipfiles:
                zipfile.ZipFile(zp).extractall(path=os.path.dirname(zp) + "/")
                if remove_zipfiles:
                    os.remove(zp)

def copy_populate_from_template(paths, info, sheet_number,zip_file):
    # Try to load group ids from the project pdf's
    all_groups = search_projects(paths, sheet_number, r"s\d{6}")
    # all_groups = get_groups_from_learn_xslx_file(paths, sheet_number)
    if len(all_groups) == 0:
        all_groups = projects_info.get_groups_from_report(repn=sheet_number-1) if sheet_number > 0 else []
    # Hopefully this did the trick and we have the groups all grouped up.

    # set up which TA approve which TA
    if any( [i['language'] not in ["en", "any"] for i in info['instructors'] ]):
        print(info['instructors'])
        raise Exception("An instructor does not have a language set. Please fix in main configuration file")
    langs = ["en", "any"]
    ifiles = get_instructor_xlsx_files(info, sheet_number)
    all_tas = {}
    for la in langs:
        tas = [i for i in info['instructors'] if i['language'] == la]
        active_tas = [t for t in tas if t['maxreports'] > 0]
        dead_tas = [t for t in tas if t['maxreports'] <= 0]

        for j, ta in enumerate(active_tas):
            sn = ta['shortname']
            all_tas[sn] = ta
            all_tas[sn]['approver'] = active_tas[(j + 1) % len(active_tas)]['shortname']
            all_tas[sn]['handins'] = []

        for j, ta in enumerate(dead_tas):
            sn = ta['shortname']
            all_tas[sn] = ta
            all_tas[sn]['approver'] = sn # dead dude approves himself
            all_tas[sn]['handins'] = []

    all_tas['transfer'] = {'maxreports': 0, 'language': 'any', 'approver': 'transfer', 'handins': []}

    for (fn, name) in ifiles:
        all_tas[name]['ifile'] = fn

    # how many groups per instructor should be preset. Larger for report 0 (group registration).
    n_groups_per_instructor = 24 + (sheet_number == 0) * 26

    if sheet_number > 0:
        # zfd = zip_file[:-4]
        # if not os.path.exists(zfd):
        #     os.mkdir(zfd)
        zfd = tempfile.TemporaryDirectory().name
        # zipfile.ZipFile(zip_file).extractall(path=tmpdir.name)

        unpack_zip_file_recursively(zip_file, destination_dir=zfd)
        # get all report handins (i.e. directories)
        ls = [l for l in glob.glob(zfd + "/*") if l[-3:] not in ["txt", "tml"]]

        handins = []
        handins_duplicated = [] # for duplicated handins. i.e. handins with existing student id's.
        protected_ids = []
        handins_MD5 = {}

        group_id_counter = 1000
        for l in ls:
            pdf_md5 = None
            lpdfs = glob.glob(l + "/*.pdf")
            try:
                group_id = int(os.path.basename(l).split("-")[2].strip().split(" ")[1])
            except Exception as e:
                group_id = group_id_counter
                group_id_counter = group_id_counter + 1

            if len(lpdfs) > 0:
                pdf = lpdfs.pop()
                try:
                    raw = parser.from_file(pdf)
                    if not 'content' in raw:
                        print("parse error; unable to parse pdf content. File is probably a bit fucky")
                        lang = "da"
                    else:
                        lang = langdetect.detect(raw['content'])
                    print(lang + ": " + pdf)
                    if lang != "en":
                        lang = "da"
                except Exception as e:
                    print("Bad encoding")
                    lang = "da"
                # Compute MD5 hash of file:
            else:
                lang = "da"

            hi = {'path': l, 'group_id': group_id, 'lang': lang, 'pdf_hash': pdf_md5}
            handins.append(hi)

        # fix error file about already handed in reports:
        used_students = []
        for h in handins:
            gid = h['group_id']
            # sid = h['student_id']
            gws = []
            for g in all_groups:
                if g['group_id'] is not None and g['group_id'] == gid:
                    # if sid in g['student_ids']:
                    gws += [s for s in g["student_ids"] if s not in protected_ids + used_students]
                    used_students += gws
            h['group'] = gws

        ta_report_langs = [(["any"], ["da"]), (["en", "any"], ["da", "en"])]

        assigned_handins = []
        for ta_l, ra_l in ta_report_langs:
            for num, h in enumerate(handins):
                if h['lang'] not in ra_l or num in assigned_handins:
                    continue
                assigned_handins.append(num)
                TA_names = [k for k in all_tas if all_tas[k]['language'] in ta_l]
                # compute number of groups per TA
                nn = [len(all_tas[n]['handins']) + 100 * (all_tas[n]['maxreports'] <= len(all_tas[n]['handins'])) for n in TA_names]
                i = np.argmin(nn)
                all_tas[TA_names[i]]['handins'].append(h)

        assert( sum([ len(all_tas[n]['handins']) for n in all_tas] ) == len(handins) )

        # Saving report assignment summary to .txt file for later reference
        summary_txt = "%s/report_%i_summary.txt"%(os.path.dirname(zip_file), sheet_number)
        with open(summary_txt, 'w') as f:
            ss = ["TA,    Approver,   number-of-reports,   group_ids,    Students"]
            for ta in all_tas:
                # for s in all_tas['niels']['students']
                handins_students = [', '.join(ha['group']) for ha in all_tas[ta]['handins'] ]
                handins_groups = ", ".join( [str(ha['group_id']) for ha in all_tas[ta]['handins'] ] )
                nha = len(handins_students)
                approver = all_tas[ta]['approver']
                # handins_groups
                ss.append(f'{ta}, {approver}, {nha},     ({handins_groups}),   ({", ".join(handins_students)})')
            f.write('\n'.join(ss))

    # write actual .xlsx files:
    template = get_template()
    for shortname in all_tas:
        ifile = all_tas[shortname]['ifile']
        corrector = all_tas[shortname]['approver']
        if sheet_number > 0:
            # Copy reports to directory (distribute amongst TAs)
            # b_dir = os.path.dirname(zip_file)
            ins_dir = "%s/project_%i_%s/"%(zfd, sheet_number, shortname)

            if not os.path.exists(ins_dir):
                os.mkdir(ins_dir)

            for handin in all_tas[shortname]['handins']:
                shutil.move(handin['path'], ins_dir)

            shutil.make_archive(os.path.dirname(zip_file) +"/"+ os.path.basename(ins_dir[:-1]), 'zip', ins_dir)
            time.sleep(2)
            print("Removing tree of reports to clear up space...")
            shutil.rmtree(ins_dir)

        if os.path.exists(ifile):
            raise Exception("File already exists")
        shutil.copyfile(template, ifile)
        wb = openpyxl.load_workbook(ifile)
        for wdex, ws in enumerate(wb.worksheets):
            if wdex != sheet_number:
                wb.remove(ws)
        ccol = 2
        sheet = wb.worksheets[0]
        if sheet_number > 0:
            sheet = write_dropdown_sumprod_sheet(sheet)

        handins_assigned_to_this_ta = all_tas[shortname]['handins']
        for i in range(len(handins_assigned_to_this_ta) + n_groups_per_instructor):
            sheet.cell(INSTRUCTOR_ROW, ccol + i).value = shortname
            if sheet_number > 0:
                if i < len(handins_assigned_to_this_ta):
                    sheet.cell(STUDENT_ID_ROW -1, ccol + i).value = handins_assigned_to_this_ta[i]['group_id']
                sheet.cell(INSTRUCTOR_ROW+1, ccol + i).value = corrector
                if i < len(handins_assigned_to_this_ta ):
                    gg = handins_assigned_to_this_ta[i]['group']

                    for j,s in enumerate(gg):
                        sheet.cell(STUDENT_ID_ROW+j, ccol+i).value = s
        wb.save(ifile)
        wb.close()
    # clean up zip file directories; since it is a tmp file, we don't have to.
    # if sheet_number > 0:
    #     zfd = zip_file[:-4]
    #     shutil.rmtree(zfd)

def write_dropdown_sumprod_sheet(sheet):
    ccol = 2
    for i in range(300):    # write 300 columns of sumprod, dropdowns. Good for courses of size up to about 800 students
        for j in range(EVALUATION_ROW_END - EVALUATION_ROW_START + 1):
            jj = j + WEIGHT_ROW_START
            min_value = sheet.cell(jj, RANGE_MIN_COL).value
            max_value = sheet.cell(jj, RANGE_MIN_COL + 1).value
            if max_value:

                rng = range(min_value, max_value + 1) if min_value >= 0 else [j for j in np.linspace(min_value, max_value, 5).flat]
                fml = '"' + ",".join([str(x) for x in rng]) + ',"'
                if min_value < 0:
                    fml = f"B{DELTA_ALLOWED_ROW}:N{DELTA_ALLOWED_ROW}"

                data_val = DataValidation(type="list", formula1=fml, allow_blank=True)
                sheet.add_data_validation(data_val)

                my_cell = sheet.cell(j + EVALUATION_ROW_START, i + ccol)
                data_val.add(my_cell)

            cl1 = get_column_letter(i + ccol)
            dfml1 = '%s%i:%s%i' % (cl1, EVALUATION_ROW_START, cl1, EVALUATION_ROW_END)
            cl2 = get_column_letter(RANGE_MIN_COL - 1)
            dfml2 = '$%s$%i:$%s$%i' % (
            cl2, WEIGHT_ROW_START, cl2, WEIGHT_ROW_START + EVALUATION_ROW_END - EVALUATION_ROW_START)
            fml2 = '=4*SUMPRODUCT(%s, %s)' % (dfml1, dfml2)
            sheet[get_column_letter(i + ccol) + str(EVALUATION_ROW_END + 1)] = fml2
    return sheet

def distribute_zip_content(info, sheet, zf_base):
    xs = get_instructor_xlsx_files(info, sheet)
    for x, TAname in xs:
        if not os.path.exists(x): continue
        if TAname == "transfer": continue
        wb = openpyxl.load_workbook(x)
        ws_x = ([wb.worksheets[0]] + [ws for ws in wb.worksheets if ws.title == "Ark1"]).pop()
        all_students = []
        for col_ins in range(1, ws_x.max_column):
            group = parse_column(ws_x, report_number=sheet, column=col_ins + 1)
            all_students += group.get('student_ids', [])

        import zipfile
        b_dir = os.path.dirname(zf_base)
        ins_dir = "%s/project_%i_%s/"%(b_dir, sheet, TAname)
        if not os.path.exists(ins_dir): os.mkdir(ins_dir)

        with zipfile.ZipFile(zf_base) as zf:
            for cfile in zf.namelist():
                for sid in all_students:
                    if cfile.startswith(sid +'/'):
                        zf.extract(cfile, ins_dir)

        shutil.make_archive(ins_dir[:-1], 'zip', ins_dir)


# Gather instructor sheets and save to main file
def gather_instructor_sheets(info):
    out = get_output_file()
    print("Gathering instructor sheets and saving them to file: ")
    print(" > %s"%out)
    template = get_template()
    shutil.copyfile(template, out)
    ts = openpyxl.load_workbook(out)

    for sheet in range(4):
        xs = get_instructor_xlsx_files(info,sheet)
        col_temp = 1

        for x, TAname in xs:
            if not os.path.exists(x): continue
            wb = openpyxl.load_workbook(x)
            ws_x = ([wb.worksheets[0]] + [ws for ws in wb.worksheets if ws.title == "Ark1"]).pop()
            tagroups = 0

            for col_ins in range(1,ws_x.max_column):
                group = parse_column(ws_x, report_number=sheet,column=col_ins+1)

                if len(group['student_ids']) > 0:
                    for r in range(ws_x.max_row):
                        if r > 50:
                            continue # don't write the part about evaluating the sheet; the TAs tend to fuck that part up.
                        dv = ws_x.cell(r+1, col_ins+1)
                        ts.worksheets[sheet].cell(r+1, col_temp+1, dv.value)
                    col_temp += 1
                    tagroups += 1

            if tagroups == 0 and not TAname == "transfer":
                print("TA: " + TAname + " sheet %i; groups found: %i" % (sheet, tagroups))
                pass

            wb.close()
        if sheet >= 1:
            write_dropdown_sumprod_sheet(ts.worksheets[sheet] )

    ts.save(out)
    ts.close()
    print("Collected xlsx instructor files. Using xlwings to load main worksheet, evaluate and save it")
    # 024
    import xlwings
    try:
        # def main() -> None:

        # info = class_information()
        # id = 's225785'
        # pid = 1
        # print(info['students'][id]['reports'][pid])

        # shutil.run( )
        # from ooodev.loader.lo import Lo
        # from ooodev.office.calc import Calc
        #
        # # fnm = sys.argv[1:]  # get file from first args
        # fnm = out
        #
        # loader = Lo.load_office(Lo.ConnectSocket(headless=True))
        # doc = Lo.open_doc(fnm=fnm, loader=loader)
        #
        # # use the Office API to manipulate doc...
        # Lo.save_doc(doc, out)  # save as a Word file
        # Lo.close_doc(doc)
        # Lo.close_office()


        book = xw.Book(out)
        book.save(out)
        book.close()
    except xlwings.XlwingsError as e:
        print("No xlsxwings installed. Sheets are not correctly evaluated.")

        # import shutil
        import subprocess
        cmd = f'cd {os.path.dirname(out)} && libreoffice --calc --headless --invisible --convert-to xlsx --outdir ../ {os.path.basename(out)}'
        output  = subprocess.run(cmd, capture_output=True, shell=True)
        time.sleep(3)
        od = os.path.dirname(os.path.dirname(out)) + "/" + os.path.basename(out)
        while not os.path.isfile(od):
            time.sleep(1)
            print("Waiting for file", od)

        shutil.move(od, out)

        print(e)


def weave_distribute_groups(info, groups, handins, shortnames):
    groups2 = []
    BG = 0
    set([g.get('instructor', "") for g in groups])
    for sid in handins:
        fg = []
        all_used_students = list(itertools.chain.from_iterable([g['student_ids'] for g in groups2]))
        for g in groups:
            if sid in g['student_ids'] and not any(set(g['student_ids'] ) & set(all_used_students)) :
                fg.append(g)
        if len(fg) > 0:
            groups2.append(fg.pop())
        else:
            groups2.append({'student_ids': [sid]})
            BG += 1
    print("Fair assigning groups. Group changes since last assignment: %i (if large -> bad TA sheet)"%BG)
    groups2 = fair_assign(info, groups2, shortnames=shortnames)
    return groups2


def _ta_maxrep_by_name(info, shortname):
    ins = [ii for ii in info['instructors'] if ii['shortname'] == shortname].pop()
    return ins['maxreports']

def fair_assign(info, groups, shortnames):
    shortnames_no_transfer = shortnames[:-1]
    groups_by_instructor = {i: [] for i in shortnames}
    n_groups = len(groups)
    MAX_groups_per_instructor = math.ceil( len(groups) / len(shortnames_no_transfer) )
    # take initial set of groups and assign them to instructors
    rem_groups = []
    for g in groups:
        found = False
        if "instructor" in g:
            i = g["instructor"].lower()
            maxreps = _ta_maxrep_by_name(info, i)
            if i in shortnames_no_transfer and len(groups_by_instructor[i]) < min([MAX_groups_per_instructor, maxreps]):
                groups_by_instructor[i].append(g)
                found = True
        if not found:
            rem_groups.append(g)
    for g in rem_groups:
        ls = [ len(groups_by_instructor[i]) if len(groups_by_instructor[i]) < _ta_maxrep_by_name(info, shortname=i) else 1000 for i in shortnames_no_transfer]
        m = np.argmin(ls)
        m = shortnames_no_transfer[m]
        groups_by_instructor[m].append(g)
    a = [len(groups_by_instructor[i]) for i in shortnames]
    for i in shortnames:
        print(i + " %i"%len(groups_by_instructor[i]))
    if sum(a) != n_groups:
        raise Exception("Group lost during fair group assignment!")
    return groups_by_instructor

def mkboxplots(students,paths):
    iscores = dict()
    for repn in range(1, 4):
        for k in students:
            s = students[k]
            g = s['reports'][repn]
            if g:
                gs = g['score']
                ins = g['instructor']
                if gs:
                    v = iscores.get(ins,[[], [], []])
                    v[repn-1].append(gs)
                    iscores[ins] = v

    NI = len(iscores.keys())
    cols = get_colors(max_colors=NI)

    def set_box_color(bp, color):
        plt.setp(bp['boxes'], color=color)
        plt.setp(bp['whiskers'], color=color)
        plt.setp(bp['caps'], color=color)
        plt.setp(bp['medians'], color=color)

    ticks = ['Report 1', 'Report 2', 'Report 3']
    plt.figure()
    bpl = []
    dw = 0.8
    lg = []
    for dex,ins in enumerate(iscores):
        data_a = iscores[ins]
        pst = np.array(range(len(data_a))) * (dw*(NI+2) ) + dw*dex
        db = plt.boxplot(data_a, positions=pst, sym='', widths=dw * 0.6/0.8)
        set_box_color(db, cols[dex])
        bpl.append(db)
        lg.append(ins)
    for dex,t in enumerate(lg):
        plt.plot([], c=cols[dex], label=t)
    plt.legend(loc='center left', bbox_to_anchor=(1, 0.5))
    plt.xlim([-1, (NI+2) * dw * 3 + 1])
    plt.xticks([ (i+0.5) *  dw * NI for i in range(len(ticks))], ticks)
    plt.ylabel("Report score")
    plt.tight_layout()
    plt.savefig(paths['instructor_project_evaluations'] + '/TA_scores.pdf')
    plt.savefig(os.path.dirname(paths['collected_project_evaluations.xlsx']) + '/TA_scores.pdf')
    plt.show()
    plt.savefig('boxcompare.png')
