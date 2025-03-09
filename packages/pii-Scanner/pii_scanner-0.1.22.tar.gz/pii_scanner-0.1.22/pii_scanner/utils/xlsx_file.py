import json
from typing import List, Dict
from openpyxl import load_workbook



def read_all_sheets(file_path: str) -> Dict[str, List[Dict[str, str]]]:
    """
    Read all sheets in an Excel (.xlsx) file and return their content as a dictionary of sheet names to lists of dictionaries.

    """
    workbook = load_workbook(filename=file_path, data_only=True)
    all_sheets_data = {}
    
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        
        # Get headers from the first row
        headers = [cell.value for cell in sheet[1]]
        
        # Extract rows into a list of dictionaries
        sheet_data = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            row_dict = {headers[i]: row[i] for i in range(len(headers))}
            sheet_data.append(row_dict)
        
        all_sheets_data[sheet_name] = sheet_data

    return all_sheets_data

def extract_column_data(data: List[Dict[str, str]], column_name: str) -> List[str]:

    return [row.get(column_name, '') for row in data]

def clean_data(data: List[str]) -> List[str]:

    return [text.strip() for text in data if text and isinstance(text, str) and text.strip()]


