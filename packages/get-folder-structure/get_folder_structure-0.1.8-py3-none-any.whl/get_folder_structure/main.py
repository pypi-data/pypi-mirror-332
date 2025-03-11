import datetime
import os
import argparse
import openpyxl

def tree(directory, exclude=None, output_txt_file="output.txt", output_xlsx_file="output.xlsx", indent="", list_type="file"):
    if exclude is None:
        exclude = []
    
    # Function to export to a text file
    def export_txt():
        with open(output_txt_file, "w", encoding="utf-8") as f:
            def walk(dir_path, level_indent):
                try:
                    entries = sorted(os.listdir(dir_path))
                    
                    for entry in entries:
                        full_path = os.path.join(dir_path, entry)
                        if entry in exclude or any(os.path.abspath(full_path).startswith(os.path.abspath(e)) for e in exclude):
                            continue
                        
                        # If 'folder' type, list only directories, if 'file' type, list both
                        if list_type == "folder" and not os.path.isdir(full_path):
                            continue
                        elif list_type == "file" and os.path.isdir(full_path):
                            # For 'file' type, include directories as well
                            f.write(f"{level_indent}{entry}/\n")
                            walk(full_path, level_indent + "    ")
                            continue
                        
                        # For 'file' type, write file entries
                        f.write(f"{level_indent}{entry}\n")
                except Exception as e:
                    print(f"Error walking directory {dir_path}: {e}")
            
            f.write(f"{directory}\n")
            walk(directory, indent)

    # Function to export to an Excel file
    def export_xlsx():
        # Create a new workbook and set up the sheet
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Directory Tree"
            
            row = 1  # Starting row for writing data
            def walk(dir_path, level_indent):
                nonlocal row
                entries = sorted(os.listdir(dir_path))
                
                for entry in entries:
                    full_path = os.path.join(dir_path, entry)
                    if entry in exclude or any(os.path.abspath(full_path).startswith(os.path.abspath(e)) for e in exclude):
                        continue
                    
                    # If 'folder' type, list only directories, if 'file' type, list both
                    if list_type == "folder" and not os.path.isdir(full_path):
                        continue
                    elif list_type == "file" and os.path.isdir(full_path):
                        # For 'file' type, include directories as well
                        path_parts = full_path.split(os.sep)
                        for col, part in enumerate(path_parts, start=1):
                            ws.cell(row=row, column=col, value=part)
                        row += 1
                        walk(full_path, level_indent + "    ")
                        continue
                    
                    # For 'file' type, write file entries
                    path_parts = full_path.split(os.sep)
                    for col, part in enumerate(path_parts, start=1):
                        ws.cell(row=row, column=col, value=part)
                    row += 1
            
            # Start walking from the root directory
            walk(directory, indent)
            
            # Save to the Excel file
            wb.save(output_xlsx_file)
        except Exception as e:
            print(f"Error exporting to Excel: {e}")

    # Export to both .txt and .xlsx
    export_txt()
    export_xlsx()

def parse_args():
    parser = argparse.ArgumentParser(description="Generate a directory tree listing.")
    parser.add_argument("-i", "--input_folder", type=str, default=".", help="Input folder to start the tree generation.")
    parser.add_argument("-e", "--exclude", type=str, nargs="*", default=["venv", "__pycache__", ".git", ".exclude", f"{datetime.datetime.now().strftime('%Y-%m-%d')}.txt", f"{datetime.datetime.now().strftime('%Y-%m-%d')}.xlsx"], help="List of files/folders to exclude.")
    parser.add_argument("-o", "--output_file", type=str, default=f"{datetime.datetime.now().strftime('%Y-%m-%d')}", help="Output file prefix for the tree structure.")
    parser.add_argument("-t", "--type", choices=["file", "folder"], default="file", help="Type of entries to list. 'folder' lists only directories, 'file' lists files and directories.")
    return parser.parse_args()

def read_exclude_file():
    exclude_file = ".exclude"
    if os.path.exists(exclude_file):
        with open(exclude_file, "r", encoding="utf-8") as f:
            return [line.strip() for line in f.readlines() if line.strip()]
    return []

def main():
    args = parse_args()
    exclude_list = args.exclude + read_exclude_file()
    output_txt_file = f"{args.output_file}.txt"
    output_xlsx_file = f"{args.output_file}.xlsx"
    tree(args.input_folder, exclude=exclude_list, output_txt_file=output_txt_file, output_xlsx_file=output_xlsx_file, list_type=args.type)
    print(f"Tree structure generated and saved to {output_txt_file} and {output_xlsx_file}")

if __name__ == "__main__":
    main()
