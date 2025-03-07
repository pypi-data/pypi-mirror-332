"""
Panacea code

Function used to do initial validation of the Excel files
"""
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter

def validate_tabs_between_spreadsheets(spreadsheet1, spreadsheet2):
    """
    Compares the sheet names between two openpyxl workbook objects to check if they are identical.

    This function compares sheet names in both workbooks, ensuring that they contain the same tabs
    (ignoring order). If there are any missing tabs in either workbook,
        it will return False and provide
    details on which sheets are missing from each spreadsheet.

    Args:
        spreadsheet1 (openpyxl.workbook.workbook.Workbook): The first workbook object to compare.
        spreadsheet2 (openpyxl.workbook.workbook.Workbook): The second workbook object to compare.

    Returns:
        dict:
            - "status": "Ok" if both workbooks have the same sheet names, "Error" otherwise.
            - "description": A general description of the comparison result.
            - "errors": A dictionary containing detailed error messages about missing tabs.
              If no errors, this will be an empty dictionary.

    Raises:
        ValueError: If either argument is not a valid openpyxl workbook object.
        InvalidFileException: If there is an issue with loading the workbook.
        Exception: For any unexpected errors during execution.
    """
    # Validate input types
    if not isinstance(spreadsheet1, Workbook) or not isinstance(spreadsheet2, Workbook):
        raise ValueError("Both arguments must be valid openpyxl workbook objects.")

    # Get sheet names from both workbooks
    sheets1 = set(spreadsheet1.sheetnames)
    sheets2 = set(spreadsheet2.sheetnames)

    # Check for missing sheets in both spreadsheets
    missing_in_1 = sheets2 - sheets1
    missing_in_2 = sheets1 - sheets2

    result = {
        "status": "Ok",
        "description": "Both spreadsheets have the same sheet names.",
        "errors": {}
    }

    if missing_in_1 or missing_in_2:
        result["status"] = "Error"
        result["description"] = "Spreadsheets have different sheet names."
        errors = {}
        if missing_in_1:
            errors["Missing In Spreadsheet 1"] = list(missing_in_1)
        if missing_in_2:
            errors["Missing In Spreadsheet 2"] = list(missing_in_2)
        result["errors"] = errors

    return result


def check_sheet_structure(sheet1, sheet2):
    """
    Compares the structure of two openpyxl worksheet objects to determine 
    if they have the same number of rows, columns, and column headers.

    This function validates whether the two worksheet objects are of the 
    correct type, checks if either sheet is empty, compares the number of 
    rows and columns, and ensures that the column headers (both name and 
    order) are the same in both sheets.

    Arguments:
        sheet1 (openpyxl.worksheet.worksheet.Worksheet): The first worksheet object to compare.
        sheet2 (openpyxl.worksheet.worksheet.Worksheet): The second worksheet object to compare.

    Returns:
        dict: A dictionary with the structure:
            - "status": "Error" or "Ok"
            - "description": The result message (either indicating success
                    or providing details on discrepancies)
            - "errors": A dictionary with error details if discrepancies are found.
              - If no discrepancies are found, errors is an empty dictionary.
    
    Example:
        sheet1 = workbook1['Sheet1']
        sheet2 = workbook2['Sheet2']
        result = check_sheet_structure(sheet1, sheet2)
        print(result)

    Notes:
        - Empty sheets are those that have no rows or columns.
        - Column comparison is case-sensitive and checks for exact matches in both name and order.
    """
    errors = {}

    # Validate input types
    if not isinstance(sheet1, Worksheet) or not isinstance(sheet2, Worksheet):
        raise ValueError("Both inputs must be valid openpyxl worksheet objects.")

    # Get the sheet names
    sheet_name1 = sheet1.title
    sheet_name2 = sheet2.title

    # Check if the sheets are empty
    if sheet1.max_row == 1 or sheet1.max_column == 1:
        errors["Empty Sheet"] = errors.get("Empty Sheet", []) + [sheet_name1]
    if sheet2.max_row == 1 or sheet2.max_column == 1:
        errors["Empty Sheet"] = errors.get("Empty Sheet", []) + [sheet_name2]

    # Check if the number of rows and columns are the same
    rows1, cols1 = sheet1.max_row, sheet1.max_column
    rows2, cols2 = sheet2.max_row, sheet2.max_column

    if (rows1, cols1) != (rows2, cols2):
        errors["Row/Column Count"] = errors.get("Row/Column Count", []) + [
            f"'{sheet_name1}' has {rows1} rows and {cols1} columns, "
            f"'{sheet_name2}' has {rows2} rows and {cols2} columns."
        ]

    # Check if the column headers are the same (both name and order)
    header1 = [sheet1.cell(row=1, column=c).value for c in range(1, cols1 + 1)]
    header2 = [sheet2.cell(row=1, column=c).value for c in range(1, cols2 + 1)]

    if header1 != header2:
        # Find out which columns are different
        diff_headers = []
        for i, (h1, h2) in enumerate(zip(header1, header2)):
            if h1 != h2:
                diff_headers.append((i + 1, h1, h2))  # Record column number and the difference
        errors["Header Mismatch"] = errors.get("Header Mismatch", []) + [
            f"Column {i}: {h1} != {h2}" for i, h1, h2 in diff_headers
        ]

    # If there are errors, return "Error" status with accumulated errors
    if errors:
        return {
            "status": "Error",
            "description": "The following discrepancies were found in the sheet structure:",
            "errors": errors
        }

    # If all checks pass, return "Ok" status
    return {
        "status": "Ok",
        "description": f"Spreadsheets '{sheet_name1}' and '{sheet_name2}' have the same structure.",
        "errors": {}
    }


def compare_formulas(sheet1, sheet2):
    """
    Compares the formulas between two openpyxl worksheet objects.

    Arguments:
        sheet1 (openpyxl.worksheet.worksheet.Worksheet): The first worksheet to compare.
        sheet2 (openpyxl.worksheet.worksheet.Worksheet): The second worksheet to compare.

    Returns:
        dict: A dictionary with status, description, and any differences:
            - If formulas are equivalent: {
                "status": "Ok",
                "description": "All formulas are equivalent",
                "errors": {}
            }
            - If formulas differ: {
                "status": "Error",
                "description": "Found formula differences",
                "errors": {
                    "Cell_Name": ["Sheet1!A1"]
                    }
                }
    """
    # Validate input types
    if not isinstance(sheet1, Worksheet) or not isinstance(sheet2, Worksheet):
        raise ValueError("Both inputs must be valid openpyxl worksheet objects.")

    # Check if the sheets have the same number of rows and columns
    rows1, cols1 = sheet1.max_row, sheet1.max_column
    rows2, cols2 = sheet2.max_row, sheet2.max_column

    if (rows1, cols1) != (rows2, cols2):
        return {
            "status": "Error",
            "description": f"Sheets have different dimensions: '{sheet1.title}' "+\
                f"has {rows1} rows & {cols1} columns, '{sheet2.title}' has "+\
                    f"{rows2} rows & {cols2} columns.",
            "errors": {}
        }

    # Dictionary to hold differing cells, grouped by their names
    differing_cells = {}

    # Compare formulas cell by cell
    for row in range(1, rows1 + 1):
        for col in range(1, cols1 + 1):
            cell1 = sheet1.cell(row=row, column=col)
            cell2 = sheet2.cell(row=row, column=col)

            # Check if both cells contain formulas (we check if cell.value starts with '=')
            if isinstance(cell1.value, str) and cell1.value.startswith('=') and \
               isinstance(cell2.value, str) and cell2.value.startswith('='):
                if cell1.value != cell2.value:
                    cell_name = f"{get_column_letter(col)}{row}"
                    # Add the differing cell to the dictionary, grouped by the cell name
                    if cell_name not in differing_cells:
                        differing_cells[cell_name] = []
                    differing_cells[cell_name].append(f"{sheet1.title}!{cell_name} "+\
                        f"({cell1.value}) != {sheet2.title}!{cell_name} ({cell2.value})")

    # If there are differences in formulas, return detailed message
    if differing_cells:
        return {
            "status": "Error",
            "description": "Found formula differences",
            "errors": differing_cells
        }

    # If all formulas are equivalent
    return {
        "status": "Ok",
        "description": "All formulas are equivalent",
        "errors": {}
    }


def check_formula_errors(sheet):
    """
    Checks for formula errors in a given openpyxl worksheet.
    
    Arguments:
        sheet (openpyxl.worksheet.worksheet.Worksheet): The worksheet to check for formula errors.
    
    Returns:
        dict: A dictionary with status, description, and any found errors in the format:
            {
                "status": "Error",
                "description": "Found errors",
                "errors": {
                    "#DIV/0!": ["Sheet1!A1"]
                }
            }
            or {"status": "Ok"} if no errors were found.
    
    Example:
        sheet = workbook['Sheet1']
        result = check_formula_errors(sheet)
        print(result)
    """
    # Validate input types
    if not isinstance(sheet, Worksheet):
        raise ValueError("Input must be valid openpyxl worksheet object.")

    error_details = {}

    # Iterate over all cells in the sheet
    for row in sheet.iter_rows():
        for cell in row:
            # Check if the cell contains an error (identified by an 'e')
            if cell.data_type == 'e':
                # If the formula's output is one of the known error strings
                if isinstance(cell.value, str):
                    cell_name = f"{sheet.title}!{get_column_letter(cell.column)}{cell.row}"
                    # Group errors by type
                    if cell.value not in error_details:
                        error_details[cell.value] = []
                    error_details[cell.value].append(cell_name)

    # If no errors were found, return the status as "Ok"
    if not error_details:
        return {"status": "Ok", "description": "No errors found", "errors": {}}

    # If errors were found, return the status as "Error" with the grouped error details
    return {
        "status": "Error",
        "description": "Found errors",
        "errors": error_details
    }
