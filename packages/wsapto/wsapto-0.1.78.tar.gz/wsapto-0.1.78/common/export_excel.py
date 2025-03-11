from django.conf import settings
import os
from datetime import datetime
from datetime import datetime, date 
import re
import json
import traceback
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.styles import Font, Alignment, Border, Side
import pandas as pd
from urllib.parse import urlparse, parse_qs
from openpyxl import Workbook
import io

def export_excel(data, headers, title_config, subtitle_config, table_start, column_widths=None, left_aligned_headers=None):
    """
    Generate an Excel file in memory with customizable content.
 
    Parameters:
    - data (list of dict): The data to populate the Excel file.
    - headers (list of str): Column headers for the Excel sheet.
    - title_config (dict): Configuration for title with 'row', 'col', and 'text'.
    - subtitle_config (list of dict): List of subtitle configurations with 'row', 'col', and 'text'.
    - table_start (dict): Dictionary with 'row' and 'col' for table start position.
    - column_widths (dict, optional): Dictionary to set custom column widths.
    - left_aligned_headers (list of str, optional): Headers that should be left-aligned.
 
    Returns:
    - BytesIO: An in-memory Excel file.
    """
    if left_aligned_headers is None:
        left_aligned_headers = []
 
    wb = Workbook()
    ws = wb.active
    ws.title = "Report"
 
    # Set Title
    title_cell = ws.cell(row=title_config['row'], column=title_config['col'], value=title_config['text'])
    title_cell.font = Font(size=20, bold=True, color="000000")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(start_row=title_config['row'], start_column=title_config['col'], 
                   end_row=title_config['row'], end_column=title_config['col'] + len(headers) - 1)
 
    # Set Subtitles
    for subtitle in subtitle_config:
        subtitle_cell = ws.cell(row=subtitle['row'], column=subtitle['col'], value=subtitle['text'])
        subtitle_cell.font = Font(size=11, italic=True, color="555555")
        alignment = subtitle.get("alignment", "center")  # Default ke center jika tidak ada
        subtitle_cell.alignment = Alignment(horizontal=alignment, vertical="center")
        ws.merge_cells(start_row=subtitle['row'], start_column=subtitle['col'], 
                    end_row=subtitle['row'], end_column=subtitle['col'] + len(headers) - 1)
 
    # Set Table Headers
    header_row_index = table_start['row']
    for col_num, header in enumerate(headers, start=table_start['col']):
        cell = ws.cell(row=header_row_index, column=col_num, value=header)
        cell.font = Font(bold=True, size=12, color="000000")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        cell.border = Border(
            top=Side(border_style="thin", color="000000"),
            left=Side(border_style="thin", color="000000"),
            right=Side(border_style="thin", color="000000"),
            bottom=Side(border_style="thin", color="000000")
        )
 
    # Populate Data Rows
    for row_index, row_data in enumerate(data, start=header_row_index + 1):
        for col_index, header in enumerate(headers, start=table_start['col']):
            cell_value = row_data.get(header, "")
            if isinstance(cell_value, (datetime, date)):
                cell_value = cell_value.strftime('%d-%m-%Y')
 
            cell = ws.cell(row=row_index, column=col_index, value=cell_value)
 
            if header in left_aligned_headers:
                cell.alignment = Alignment(horizontal="left", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center")
 
            cell.border = Border(
                top=Side(border_style="thin", color="000000"),
                left=Side(border_style="thin", color="000000"),
                right=Side(border_style="thin", color="000000"),
                bottom=Side(border_style="thin", color="000000")
            )
 
    # Adjust Column Widths
    for col_num, header in enumerate(headers, start=table_start['col']):
        if column_widths and header in column_widths:
            ws.column_dimensions[ws.cell(row=table_start['row'], column=col_num).column_letter].width = column_widths[header]
        else:
            max_length = max(len(str(header)), *(len(str(row_data.get(header, ""))) for row_data in data))
            ws.column_dimensions[ws.cell(row=table_start['row'], column=col_num).column_letter].width = max_length + 5
 
    # Save workbook to an in-memory buffer
    output_buffer = io.BytesIO()
    wb.save(output_buffer)
    output_buffer.seek(0)
    return output_buffer