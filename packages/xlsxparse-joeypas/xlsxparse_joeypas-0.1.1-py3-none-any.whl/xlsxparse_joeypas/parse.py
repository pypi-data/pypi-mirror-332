from openpyxl import load_workbook
from openpyxl.workbook.external_link import ExternalLink
from openpyxl.workbook.defined_name import DefinedNameDict
from openpyxl.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.worksheet.cell_range import CellRange
from pathlib import PureWindowsPath
from urllib.parse import unquote
import re

def build_defined_ranges(defined_names: DefinedNameDict) -> [dict[str, str, CellRange]]:
    """Create a dictionary for each defined name in the Workbook.

    Returned dictionary definition:
    name  -- the defined name
    sheet -- sheet the name belongs to
    range -- CellRange object specifing the range of cells the name applies to
    """
    defined_names = [x for x in defined_names.items() if not x[0].startswith("_xlchart")]
    ret = []
    for name in defined_names:
        defn = name[0]
        for title, coord in name[1].destinations:
            ret.append({
                'name': defn,
                'sheet': title,
                'range': CellRange(coord),
            })
    return ret

def is_defined(cell: Cell, defined_ranges: [dict[str, str, CellRange]]) -> [str]:
    """Finds if a cell belongs to a defined name.

    If the cell belongs to one or more names, the function
    returns a list of the names. Otherwise an empty list is
    returned.

    defined_ranges argumnet must be a list of dictionaries returned
    from build_defined_ranges.
    """
    sheet_name = cell.parent.title
    coord = cell.coordinate
    ranges = [x['name'] for x in defined_ranges if x['sheet'] == sheet_name and not x['range'].isdisjoint(CellRange(coord))]
    return ranges

def extract_references(formula: str, curr_sheet: Worksheet, links: [ExternalLink]) -> [dict[str, str, str]]:
    """Parse an excel formula for refrences.

    Returns any cells refrenced in the formula, along with
    what sheet, and what file the cells belong to.
    """
    ref_pattern = r"(?:(?: *'?\[([^\]]+)\])?([^'=,\[]+)'?\!)?([A-Z|$]+\d+(?::[A-Z|$]+\d+)?)"
    matches = re.findall(ref_pattern, formula)
    refs = []

    for match in matches:
        ref = {}
        workbook, sheet, cell = match
        if (workbook):
            ref["file"] = PureWindowsPath(unquote(links[int(workbook)-1].file_link.Target)).name
        if (sheet):
            ref["sheet"] = sheet
        else:
            ref["sheet"] = curr_sheet.title
        ref["cell"] = cell
        refs.append(ref)

    return refs

def get_names(cell: Cell, defined_ranges: [dict[str, str, CellRange]] = []):
    """Attempts to find labels for the provided cell.

    If defined_ranges list is provided, the function will
    try to determine labels using is_defined. Otherwise 
    searches backwards by row/col.
    """
    if len(defined_ranges) > 0:
        defd = is_defined(cell, defined_ranges)
        if (len(defd) > 0):
            return tuple(defd)


    sheet = cell.parent
    row_label = None
    col_label = None
    i = cell.row
    j = cell.col_idx
    while (j > 1):
        if sheet.cell(i, j).data_type == 's' or sheet.cell(i, j).data_type == 'd':
            row_label = str(sheet.cell(i, j).value)
            break
        j -= 1
    j = cell.col_idx
    while (i > 1):
        if sheet.cell(i, j).data_type == 's' or sheet.cell(i, j).data_type == 'd':
            col_label = str(sheet.cell(i, j).value)
            break
        i -= 1
    return (row_label, col_label)

def parse_excel_formulas(
    sheet: Worksheet, 
    links: [ExternalLink], 
    defined_ranges: [dict[str, str, CellRange]] = []
) -> dict[dict[str, str, dict]]:
    """Find all refrences in a given sheet"""
    formulas = {}

    for row in sheet.iter_rows():
        for cell in row:
            if (isinstance(cell.value, str) and cell.value.startswith('=')):
                names = get_names(cell, defined_ranges)
                formulas[cell.coordinate] = {
                    "names": names,
                    "formula": cell.value,
                    "references": extract_references(cell.value, sheet, links),
                }

    return formulas

def parse_all_sheets(file_path: str) -> dict[dict]:
    wb = load_workbook(file_path, data_only=False)
    defined_ranges = build_defined_ranges(wb.defined_names)
    all_refs = {}

    for sheet in wb.sheetnames:
        all_refs[sheet] = {
            "items": parse_excel_formulas(wb.get_sheet_by_name(sheet), wb._external_links, defined_ranges),
        }

    return all_refs

def parse_single_sheet(file_path: str, sheet_name: str) -> dict[dict]:
    wb = load_workbook(file_path, data_only=False)
    sheet = wb[sheet_name]
    defined_ranges = build_defined_ranges(wb.defined_names)


    return parse_excel_formulas(sheet, wb._external_links, defined_ranges)
