import sys
import os
from PySide6.QtWidgets import QApplication, QFileDialog, QMessageBox
from PySide6.QtWidgets import QMainWindow
from openpyxl import load_workbook
from email_splitter_ui import Ui_MainWindow  # Import the UI generated from Qt Designer


class EmailSplitterApp(QMainWindow):
    def __init__(self):
        super().__init__()
        self.ui = Ui_MainWindow()
        self.ui.setupUi(self)

        # Connect buttons to functions
        self.ui.btnSelectInput.clicked.connect(self.select_input_file)
        self.ui.btnSelectOutput.clicked.connect(self.select_output_file)
        self.ui.btnProcess.clicked.connect(self.process_file)

        # Default values
        self.input_file = ""
        self.output_file = ""

    def select_input_file(self):
        file_path, _ = QFileDialog.getOpenFileName(self, "Select Input Excel File", "", "Excel Files (*.xlsx)")
        if file_path:
            self.input_file = file_path
            self.ui.lblInputFile.setText(os.path.basename(file_path))

    def select_output_file(self):
        file_path, _ = QFileDialog.getSaveFileName(self, "Select Output Excel File", "", "Excel Files (*.xlsx)")
        if file_path:
            self.output_file = file_path
            self.ui.lblOutputFile.setText(os.path.basename(file_path))

    def process_file(self):
        if not self.input_file or not self.output_file:
            QMessageBox.warning(self, "Error", "Please select both input and output files.")
            return

        sheet_name = self.ui.txtSheetName.text().strip()
        email_column = self.ui.txtEmailColumn.text().strip()

        if not sheet_name or not email_column:
            QMessageBox.warning(self, "Error", "Please enter sheet name and email column.")
            return

        try:
            self.split_emails_in_excel(self.input_file, self.output_file, sheet_name, email_column)
            QMessageBox.information(self, "Success", "File processed successfully!")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"An error occurred:\n{str(e)}")

    def split_emails_in_excel(self, input_file, output_file, sheet_name, email_column):
        wb = load_workbook(input_file)
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found in the Excel file.")

        ws = wb[sheet_name]

        # Extract headers
        headers = [cell.value for cell in ws[1]]
        if email_column not in headers:
            raise ValueError(f"Column '{email_column}' not found in '{sheet_name}'.")

        email_col_idx = headers.index(email_column) + 1

        new_data = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            row_list = list(row)
            email_values = str(row_list[email_col_idx - 1]).split(',')
            for email in email_values:
                new_row = row_list.copy()
                new_row[email_col_idx - 1] = email.strip()
                new_data.append(new_row)

        # Clear the original sheet (except header)
        ws.delete_rows(2, ws.max_row)

        # Write new data
        for i, row_data in enumerate(new_data, start=2):
            for j, value in enumerate(row_data, start=1):
                ws.cell(row=i, column=j, value=value)

        wb.save(output_file)


def main():
    app = QApplication(sys.argv)
    window = EmailSplitterApp()
    window.show()
    app.exec()


if __name__ == "__main__":
    main()
